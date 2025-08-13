import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
import subprocess
import logging
from logger_config import setup_logger
import os
import glob

# Inicializa o logger
logger = setup_logger()

# Configurações iniciais
IMAGES_PATH = "img"
PASTA_PLANILHAS = r"\\192.168.1.250\Programas\Quick Books\Robo Quickboks"

def obter_ultima_planilha(pasta):
    """
    Encontra a planilha Excel mais recente na pasta especificada
    """
    try:
        # Busca todos os arquivos .xlsx na pasta
        padrao = os.path.join(pasta, "*.xlsx")
        arquivos_excel = glob.glob(padrao)
        
        if not arquivos_excel:
            logger.error(f"Nenhum arquivo Excel encontrado na pasta: {pasta}")
            return None
        
        # Ordena por data de modificação (mais recente primeiro)
        arquivo_mais_recente = max(arquivos_excel, key=os.path.getmtime)
        
        logger.info(f"Arquivo mais recente encontrado: {arquivo_mais_recente}")
        return arquivo_mais_recente
        
    except Exception as e:
        logger.error(f"Erro ao buscar a última planilha: {e}")
        return None

def process_planilha():
    try:
        logger.info("Iniciando o processamento da planilha.")

        # Obter a planilha mais recente da pasta
        CAMINHO_PLANILHA_ORIGINAL = obter_ultima_planilha(PASTA_PLANILHAS)
        
        if not CAMINHO_PLANILHA_ORIGINAL:
            logger.error("Não foi possível encontrar uma planilha para processar.")
            return
        
        # Verificar se o arquivo existe
        if not os.path.exists(CAMINHO_PLANILHA_ORIGINAL):
            logger.error(f"Arquivo não encontrado: {CAMINHO_PLANILHA_ORIGINAL}")
            return
            
        CAMINHO_NOVA_PLANILHA = r"C:\\Users\\samuca2\\Documents\\base_robo_quickboks\\nova_planilha.xlsx"

        logger.debug(f"Caminho da planilha original: {CAMINHO_PLANILHA_ORIGINAL}")
        logger.debug(f"Caminho da nova planilha: {CAMINHO_NOVA_PLANILHA}")

        # Abrir a planilha original
        logger.info("Abrindo a planilha original.")
        wb_original = openpyxl.load_workbook(CAMINHO_PLANILHA_ORIGINAL, data_only=True)
        
        # Verificar se a aba existe
        if "VENDAS POR CLIENTE" not in wb_original.sheetnames:
            logger.error(f"Aba 'VENDAS POR CLIENTE' não encontrada. Abas disponíveis: {wb_original.sheetnames}")
            return
            
        sheet_original = wb_original["VENDAS POR CLIENTE"]
        logger.debug("Planilha original carregada com sucesso.")

        # Identificar os cabeçalhos na linha 4
        linha_cabecalho = 4
        cabecalhos = ["LW", "HD", "WM", "WF", "TS", "RD", "CS"]
        colunas_cabecalhos = []

        logger.info(f"Identificando os cabeçalhos na linha {linha_cabecalho}.")
        for col in range(1, sheet_original.max_column + 1):
            valor_celula = sheet_original.cell(row=linha_cabecalho, column=col).value
            if valor_celula in cabecalhos:
                colunas_cabecalhos.append(col)
                logger.debug(f"Encontrado cabeçalho '{valor_celula}' na coluna {col}.")

        if len(colunas_cabecalhos) < 2:
            logger.error("Não foi possível encontrar ao menos duas estruturas completas.")
            raise ValueError("Não há pelo menos duas estruturas completas para selecionar a penúltima.")

        logger.info(f"Colunas dos cabeçalhos encontrados: {colunas_cabecalhos}")

        # Determinar o intervalo completo da **penúltima estrutura**
        penultima_estrutura_inicio = colunas_cabecalhos[-len(cabecalhos) * 2]  # Seleciona a penúltima estrutura
        penultima_estrutura_fim = penultima_estrutura_inicio + len(cabecalhos) * 2 - 1
        logger.debug(f"Intervalo da penúltima estrutura: início={penultima_estrutura_inicio}, fim={penultima_estrutura_fim}")

        # Copiar os dados abaixo da penúltima estrutura, ignorando linhas vazias
        logger.info("Copiando os dados abaixo da penúltima estrutura, ignorando linhas vazias.")
        dados_copiados = []
        coluna_a_dados = []

        for row in sheet_original.iter_rows(min_row=linha_cabecalho + 1, max_row=sheet_original.max_row,
                                            min_col=1, max_col=penultima_estrutura_fim):
            linha_dados = [cell.value for cell in row]
            if any(linha_dados):  # Ignorar linhas vazias
                coluna_a_dados.append(str(row[0].value) if row[0].value is not None else "")
                dados_copiados.append(linha_dados[penultima_estrutura_inicio - 1:penultima_estrutura_fim])
            else:
                logger.debug("Linha ignorada por estar vazia ou conter apenas valores None.")

        logger.debug(f"Número de linhas copiadas: {len(dados_copiados)}")

        # Criar a nova planilha e colar os dados
        wb_nova = Workbook()
        sheet_nova = wb_nova.active
        sheet_nova.title = "Estrutura Copiada"
        logger.info("Nova planilha criada com sucesso.")

        # Adicionar os dados copiados na nova planilha
        cores = ["FF9999", "99FF99", "9999FF", "FFFF99", "FF99FF", "99FFFF", "FFCC99"]

        logger.info("Adicionando dados copiados na nova planilha.")
        for i, (coluna_a, row_data) in enumerate(zip(coluna_a_dados, dados_copiados), start=1):
            sheet_nova.cell(row=i, column=1).value = coluna_a
            col_offset = 2
            for j, value in enumerate(row_data):
                sheet_nova.cell(row=i, column=col_offset).value = value
                col_offset += 1
                if (j + 1) % 2 == 0:
                    col_offset += 1

        # Aplicar cores aos cabeçalhos e colunas correspondentes
        logger.info("Aplicando cores aos cabeçalhos e colunas correspondentes.")
        for idx, col_start in enumerate(range(2, len(cabecalhos) * 3 + 2, 3)):
            fill = PatternFill(start_color=cores[idx % len(cores)], end_color=cores[idx % len(cores)], fill_type="solid")
            for row in range(1, len(dados_copiados) + 1):
                sheet_nova.cell(row=row, column=col_start).fill = fill
                sheet_nova.cell(row=row, column=col_start + 1).fill = fill

        # Salvar a nova planilha
        logger.info(f"Salvando a nova planilha em {CAMINHO_NOVA_PLANILHA}.")
        wb_nova.save(CAMINHO_NOVA_PLANILHA)
        logger.info(f"Dados copiados para a nova planilha: {CAMINHO_NOVA_PLANILHA}")

        # Aguardar antes de chamar o próximo script
        logger.info("Aguardando 5 segundos antes de chamar o próximo script.")
        time.sleep(5)

        # Chamar o arquivo 4-lw.py
        logger.info("Executando o script '4-lw.py'...")
        subprocess.run(["python", "4-lw.py"], check=True)
        logger.info("Script '4-lw.py' executado com sucesso.")

    except Exception as e:
        logger.exception(f"Erro durante o processamento da planilha: {e}")

if __name__ == "__main__":
    process_planilha()